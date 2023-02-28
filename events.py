"""
This project intends to read the required files from Google Drive, and 
then based on the naming convention applied to those files, analyse and 
categorize the available data into the necessary format. 

Project Creation: Feb 3, 2023
Lead Developer: Sam Alex Koshy
Project Contributors: Ashok Immanuel, Rohini V.
"""


from __future__ import print_function

import io
import os
import os.path
from json import dumps
from pprint import PrettyPrinter
import sys

# Install necessary libraries with pip install -r requirements.txt
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# If modifying these scopes, delete the file token.json.
SCOPES = [
    "https://www.googleapis.com/auth/drive"
]

pp = PrettyPrinter(indent=4)


class ExcelWorker():
    """The class that will handle interaction with the excel workbooks."""

    def __init__(self) -> None:
        """Initialize the class"""
        self.classification: dict[str, dict[str, str]] = {}
        self.code_groups:dict[str, str] = {}
        self.read_classification_exl()

    def read_classification_exl(self):
        """Read the classification categories."""
        # Load the workbook, if not exit the program.
        try:
            self.workbook:Workbook = load_workbook("doc_classification.xlsx")
        except FileNotFoundError: 
            print("Classification file not found.")
            sys.exit()

        worksheet: Worksheet
        # Loop through all sheets and categorize each code.
        for worksheet in self.workbook:
            if worksheet.title == "NAAC Quantitative":
                self.final_codes: dict[str, list[str]] = {}
                self.categories: dict[int, str] = {}
                self.spec_list: dict[str, str] = {}

                for row in worksheet.iter_rows(min_col=1, max_col=4, 
                                               min_row=4, values_only=True):
                    category:str = row[0] or category
                    spec:str = row[1] or spec
                    letter_code, full_form = row[2], row[3]

                    if category and int(spec[0]) not in self.categories:
                        self.categories[int(spec[0])] = category
                    # print(spec, letter_code, full_form) 
                    if letter_code is not None:
                        if letter_code in self.final_codes:
                            self.final_codes[letter_code].append(spec)
                        else:
                            self.final_codes[letter_code] = [spec]
                    if spec and spec not in self.spec_list:
                        self.spec_list[spec] = category
                # print(self.final_codes)
                # print(self.spec_list)
                # print(self.categories)

            else:
                ws_dict = {}
                for row in worksheet.iter_rows(min_col=2, max_col=3):
                    code, name = row[0].value, row[1].value
                    if code is not None and name is not None:
                        ws_dict[code] = name
                        self.code_groups[code] = str(worksheet.title)
                if ws_dict != {}:
                    self.classification.update({worksheet.title: ws_dict})
        # with open("classification.json", "w") as file:
        #     class_obj = dumps(self.classification, indent=4)
        #     file.write(class_obj)
        return self.classification

    def write_to_excel(self, drive_data: dict[str, dict[str, dict[str, int]]], 
                       exempted: list[tuple[str, str]]):
        """Write data from the drive to the excel sheet."""
        workbook = Workbook()
        workbook.active.title = "exempted"

        # Loop through all the categories to create new sheets.
        for category in drive_data:
            worksheet: Worksheet = workbook.create_sheet(category, -1)
            category_data = drive_data[category]
            worksheet.append(["YEAR", "CLASSIFICATION", "COUNT"])
            for i in range(1, 4):
                worksheet[f"{get_column_letter(i)}1"].alignment = Alignment(horizontal="center")
                worksheet[f"{get_column_letter(i)}1"].font = Font(bold=True, size=12)
            start, stop, width = 2, 2, 16
            for year in category_data:
                worksheet[f"A{start}"] = year
                year_data = category_data[year]
                for code in year_data:
                    name = self.classification[category][code]
                    worksheet[f"B{stop}"] = name
                    worksheet[f"C{stop}"] = year_data[code]
                    width = max(width, len(name))
                    stop += 1

                # Merge the cells of same years, and center the alignment.
                worksheet.merge_cells(f"A{start}:A{stop-1}")
                worksheet[f"A{start}"].alignment = Alignment(horizontal="center", 
                                                             vertical="center")
                start = stop
            worksheet.column_dimensions["B"].width = width
            worksheet.column_dimensions["A"].width = 13
        worksheet = workbook["exempted"]
        worksheet.append(["File Name", "Folder name"])
        width1, width2 = 13, 13
        for value in exempted:
            worksheet.append(value)
            width1, width2 = max(width1, len(value[0])), max(width2, len(value[1]))
        worksheet.column_dimensions["A"].width = width1
        worksheet.column_dimensions["B"].width = width2
        while True:
            try:
                # try:
                #     os.system("taskkill/im EXCEL.EXE categorized.xlsx")
                # except:
                #     pass
                workbook.save("categorized.xlsx")
            except PermissionError:
                try:
                    os.system("taskkill/im EXCEL.EXE categorized.xlsx")
                except:
                    pass
            else:
                break
        # os.system("start EXCEL.EXE categorized.xlsx")

    def write_to_naac(self, drive_data: dict[str, dict[str, dict[str, int]]]):
        """Write data to excel sheet in naac required format."""
        spec_data = {}
        for category_data in drive_data.values():
            for year in category_data:
                if year == "2022-2023":
                    year_data = category_data[year]
                    for code in year_data:
                        if self.final_codes.get(code):
                            for spec in self.final_codes.get(code):
                                spec_data.update({spec: spec_data.get(spec, 0) + year_data[code]})
        order_list = sorted(spec_data.keys())
        spec_data = dict([(i, spec_data[i]) for i in order_list])
        naac_wb = Workbook()
        naac_ws: Worksheet = naac_wb.active
        start, stop = 1, 1
        old_number = 0
        width = 13
        # print(spec_data)
    
        # * Entering the data that is there.
        # for spec in spec_data:
        #     number = int(spec[0])
        #     if old_number != number:
        #         naac_ws.merge_cells(f"A{start}:A{stop}")
        #         start = stop + 1
        #         naac_ws[f"A{start}"].alignment = Alignment(horizontal="center", 
        #                                                    vertical="center")
        #         old_number = number
        #         # stop += 1
        #         naac_ws[f"A{start}"] = self.categories[number]
        #         width = max(width, len(self.categories[number])*1.3)
        #     stop += 1
        #     naac_ws[f"B{stop}"] = spec
        #     naac_ws[f"C{stop}"] = spec_data[spec]
        # naac_ws.column_dimensions["A"].width = width
    
        # * Entering all specification codes.
        naac_ws.append(["Classification", "Code", "Count"])
        start, word = 1, None
        for num, spec in enumerate(self.spec_list, 2):
            naac_ws.append([self.spec_list[spec], spec, spec_data.get(spec, 0)])
            width = max(width, len(self.spec_list[spec])*1.2)
            if word != self.spec_list[spec]:
                naac_ws.merge_cells(f"A{start}:A{num-1}")
                naac_ws[f"A{start}"].alignment = Alignment(horizontal="center", 
                                                           vertical="center")
                start, word = num, self.spec_list[spec]
        else:
            naac_ws.merge_cells(f"A{start}:A{num}")
            naac_ws[f"A{start}"].alignment = Alignment(horizontal="center", 
                                                        vertical="center")
            for i in range(1, 4):
                naac_ws[f"{get_column_letter(i)}1"].alignment = Alignment(
                    horizontal="center", vertical="center")
                naac_ws[f"{get_column_letter(i)}1"].font = Font(bold=True, size=12)
            naac_ws.column_dimensions["A"].width = width

        while True:
            try:
                naac_wb.save("naac.xlsx")
                break
            except PermissionError:
                print("Failed to save naac.xlsx")
                os.system("taskkill /im EXCEL.EXE naac.xlsx")
        os.system("start EXCEL.EXE naac.xlsx")


class DriveReader():
    """This project aims to read files in a drive and categorize them."""

    def __init__(self) -> None:
        """Initialize the class."""
        self.creds = None
        self.validate_user()
        # self.download_sheet()

        # Create an object of excel class.
        self.excelWorker = ExcelWorker()
        self.categories = self.excelWorker.classification
        self.code_keys = self.excelWorker.code_groups

    def validate_user(self):
        """Validate the program if the user who runs it is registered."""
        # The file stores user's access and refresh tokens, and is created 
        # automatically when first authorization flow is completed.
        if os.path.exists("token.json"):
            self.creds = Credentials.from_authorized_user_file("token.json", 
                                                               SCOPES)

        if not self.creds or not self.creds.valid:
            if self.creds and self.creds.expired and self.creds.refresh_token:
                self.creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    "credentials.json", SCOPES)
                self.creds = flow.run_local_server(port=0)
            # Save the credentials for the next run.
            with open("token.json", "w") as token:
                token.write(self.creds.to_json())

        # Create a connection with drive.
        self.service = build("drive", "v3", credentials=self.creds)

        # Set data empty and then start execution.
        self.data:dict[str, dict[str, dict[str, int]]] = {}
        self.exempt:list[tuple(str, str)] = []

    def sort_files_in_folder(self, category_name: str):
        """Sort the files in the folder."""
        folders = self.search_folder(category_name)
        if folders is None or folders == []:
            return
        for folder in folders:
            folder_id, folder_name = folder.get("id"), folder.get("name")
            try:
                page_token = None
                while True:
                    # Search for all files with the folder as parent.
                    response = self.service.files().list(
                        q=f"'{folder_id}' in parents and trashed = false",
                        spaces='drive',
                        fields='nextPageToken, files(name)'
                    ).execute()

                    for file in response.get("files"):
                        # print(file)
                        if file.get("name") is not None:
                            if_failed = self.classify_file(file.get("name"))
                            if if_failed is not None:
                                self.exempt.append((if_failed, folder_name))
                    page_token = response.get("nextPageToken", None)

                    if page_token is None:
                        break

            except HttpError as error:
                print(f"An error occurred: {error}")


    def search_folder(self, category_name: str):
        """Search for a specific folder."""
        try:
            response = self.service.files().list(
                q=f"name contains '{category_name}' and mimeType = \
                    'application/vnd.google-apps.folder'"
            ).execute()
            return response.get("files", None)

        except HttpError as error:
            print(f"An error occurred: {error}")
            return None

    def search_file(self, file_name: str):
        """Search for a specific folder."""
        try:
            response = self.service.files().list(
                q=f"name contains '{file_name}'"
            ).execute()
            return response.get("files", None)

        except HttpError as error:
            print(f"An error occurred: {error}")
            return None

    def classify_file(self, name:str):
        """Classify the file in categories based on naming structure."""
        try:
            date, code, extra = name.split("_", 2)
            code = code.upper()
            if code not in self.code_keys:
                raise KeyError
        except ValueError:
            return name
        except KeyError:
            return name
        else:
            try:
                year, month = int(date[:4]), int(date[4:6])
                # print(year, month, day)
                if month > 0 and month < 5:
                    year = f"{year-1}-{year}"
                else:
                    year = f"{year}-{year+1}"
            except ValueError:
                return name
            else:
                category = self.code_keys[code]
                category_data = self.data.get(category, {"0": {"0": 0}})
                year_data = category_data.get(year, {"0": 0})
                code_val = year_data.get(code, 0)
                year_data.update({code: code_val + 1})
                if "0" in year_data:
                    year_data.pop("0")
                category_data.update({year: year_data})
                if "0" in category_data: 
                    category_data.pop("0")
                self.data.update({category: category_data})
                if "0" in self.data:
                    self.data.pop("0")

    def download_sheet(self):
        """Download the required excel sheet."""
        self.excel_sheet_id = "1b5yJfOIWCHXdr7VbFxoNLs_SI5zPR7CL0MsCI1zqWaM"
        try:
            # response = self.service.files().export(fileId=self.excel_sheet_id,
            #                                        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet").execute()
            # print(response)
            # print(dir(response))
            request = self.service.files().export_media(fileId=self.excel_sheet_id,
                                               mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file = io.BytesIO()
            downloader = MediaIoBaseDownload(file, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()

            with open("doc_classification.xlsx", "wb") as write_file:
                write_file.write(file.getbuffer())

            print(F'Download {int(status.progress() * 100)}.')

        except HttpError as error:
            print(f"{error} has occurred.")

    def main(self):
        """The main function of the class."""
        folders = [
            "Alumni",
            "Curriculum",
            "Events",
            "Faculty",
            "Mou_Consultancy",
            "Research"
        ]
        for folder in folders:
            self.sort_files_in_folder(folder)
        else:
            for category in self.data:
                for year in self.data[category]:
                    year_data = self.data[category][year]
                    order_list = sorted(year_data.keys())
                    new_list = dict([(i, year_data[i]) for i in order_list])
                    self.data[category][year] = new_list
            # with open("data.json", "w") as file:
            #     data_obj = dumps(self.data, indent=4)
            #     file.write(data_obj)
            # with open("exempt.json", "w") as file:
            #     exempt_obj = dumps(self.exempt, indent=4)
            #     file.write(exempt_obj)
        # self.excelWorker.write_to_excel(self.data, self.exempt)
        self.excelWorker.write_to_naac(self.data)
        # self.classify_file("20220730_cprs_rv_1.pdf")
        # pp.pprint(self.search_file("doc_classification"))


if __name__ == "__main__":
    # Driver Code
    try:
        DR = DriveReader()
        if DR.creds and DR.creds.valid:
            DR.main()
        else:
            sys.exit()
    except KeyboardInterrupt:
        print("\n\nExiting program by interrupt.")