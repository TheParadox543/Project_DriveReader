from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# wb = Workbook()
# ws: Worksheet = wb.active
# ws.column_dimensions: ColumnDimension
# for i in range(1, 20):
#     ws.cell(row=1, column=i)
#     ws.column_dimensions[get_column_letter(i)].width = i
# wb.save("trial.xlsx")

wb = load_workbook("categorized.xlsx")
ws = wb["RESEARCH"]
for cell in ws["B"]:
    print(len(cell.value))
print(ws.column_dimensions["C"].width)