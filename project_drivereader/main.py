"""
This project intends to read the required files from Google Drive, and 
then based on the naming convention applied to those files, analyse and 
categorize the available data into the necessary format. 

Project Creation: Feb 3, 2023
Lead Developer: Sam Alex Koshy
Project Contributors: Ashok Immanuel, Rohini V.
"""


# Install necessary libraries with pip install -r requirements.txt
from __future__ import print_function

from io import BytesIO
from json import dumps, load
from os import path, system as ossystem
from pprint import PrettyPrinter
from sys import exit as sysexit

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# If modifying these scopes, delete the file token.json.
SCOPES = [
    "https://www.googleapis.com/auth/drive"
]

pp = PrettyPrinter(indent=4)


class Category():
    """"""

class Classification():
    """"""

class Code():
    """"""

class Name():
    """"""

class Year():
    """"""


class ExcelWorker():
    """The class that will handle interaction with the excel workbooks."""

    def __init__(self) -> None:
        """Initialize the class."""
        self.read_classification_exl()

    def read_classification_exl(self):
        """Read the classification categories."""
        # Load the workbook, if not exit the program.
        try:
            self.doc_wb:Workbook = load_workbook("data/doc_classification.xlsx")
        except FileNotFoundError: 
            print("Classification file not found.")
            sysexit()

        # * The following are the data to be extracted from the excel sheet.
        # Category - The name of major types of requirements.
        # Classification - The numeric code.
        # Code - A 4 letter (sometimes 3) that represents what it is.
        # Name - The full name of the previous mentioned.

        # "RPIF": [full name, category, [1.2, 2,3]]
        self.code_list: dict[Code, tuple[Name, Category, list[Classification]]] = {} 
        self.classification_list: dict[Classification, Category] = {} # "2.3.4": "Research"
        ws: Worksheet

        # Loop through all sheets and categorize each code.
        for ws in self.doc_wb:
            if ws.title == "NAAC Quantitative":

                for row in ws.iter_rows(min_col=1, max_col=4, 
                                               min_row=3, values_only=True):
                    category: Category = row[0] or category
                    classification: Classification = row[1] or classification
                    code: Code = row[2]
                    name: Name = row[3]

                    if code and name:
                        if code not in self.code_list:
                            self.code_list[code] = [name, category,
                                                     [classification]]
                        else:
                            self.code_list[code][2].append(classification)
                    if classification not in self.classification_list:
                        self.classification_list[classification] = category

            else:
                for row in ws.iter_rows(min_col=2, max_col=3):
                    code, name = row[0].value, row[1].value
                    if code and name and code not in self.code_list:
                        self.code_list[code] = [name, category, ["Unknown"]]
        with open("data/code_list.json", "w") as file:
            code_obj = dumps(self.code_list, indent=4)
            file.write(code_obj)

    def write_data_to_excel(self, drive_data: dict[Category, dict[Year, dict[Code, int]]], 
                       exempted: list[tuple[Name, str]]):
        """Write data from the drive to the excel sheet."""
        workbook = Workbook()
        workbook.active.title = "exempted"

        # Loop through all the categories to create new sheets.
        for category in drive_data:
            worksheet: Worksheet = workbook.create_sheet(category, -1)
            category_data = drive_data[category]
            worksheet.append(["YEAR", "CLASSIFICATION", "COUNT"])
            for i in range(1, 4):
                worksheet[f"{get_column_letter(i)}1"].alignment = Alignment(horizontal="center")
                worksheet[f"{get_column_letter(i)}1"].font = Font(bold=True, size=12)
            start, stop, width = 2, 2, 16
            for year in category_data:
                worksheet[f"A{start}"] = year
                year_data = category_data[year]
                for code in year_data:
                    name = self.code_list[code][0]
                    worksheet[f"B{stop}"] = name
                    worksheet[f"C{stop}"] = year_data[code]
                    width = max(width, len(name))
                    stop += 1

                # Merge the cells of same years, and center the alignment.
                worksheet.merge_cells(f"A{start}:A{stop-1}")
                worksheet[f"A{start}"].alignment = Alignment(horizontal="center", 
                                                             vertical="center")
                start = stop
            worksheet.column_dimensions["B"].width = width
            worksheet.column_dimensions["A"].width = 13
        worksheet = workbook["exempted"]
        worksheet.append(["File Name", "Folder name"])
        width1, width2 = 13, 13
        for value in exempted:
            worksheet.append(value)
            width1, width2 = max(width1, len(value[0])), max(width2, len(value[1]))
        worksheet.column_dimensions["A"].width = width1
        worksheet.column_dimensions["B"].width = width2
        while True:
            try:
                # try:
                #     os.system("taskkill/im EXCEL.EXE categorized.xlsx")
                # except:
                #     pass
                workbook.save("data/categorized.xlsx")
            except PermissionError:
                try:
                    ossystem("taskkill/im EXCEL.EXE categorized.xlsx")
                except:
                    pass
            else:
                break
        # os.system("start EXCEL.EXE categorized.xlsx")

    def write_naac_data_to_excel(self, 
                                 drive_data: dict[Category, dict[Year, dict[Code, int]]]):
        """Write data to excel sheet in naac required format."""
        spec_data = {}
        for category_data in drive_data.values():
            for year, year_data in category_data.items():
                if year == "2022-2023":
                    for code, value in year_data.items():
                        if self.code_list.get(code):
                            for spec in self.code_list[code][2]:
                                spec_data.update({spec: spec_data.get(spec, 0) + value})
        order_list = sorted(spec_data.keys())
        spec_data = dict([(i, spec_data[i]) for i in order_list])
        naac_wb = Workbook()
        naac_ws: Worksheet = naac_wb.active
        start, stop = 1, 1
        old_number = 0
        width = 13
        # print(spec_data)
    
        # * Entering the data that is there.
        # for spec in spec_data:
        #     number = int(spec[0])
        #     if old_number != number:
        #         naac_ws.merge_cells(f"A{start}:A{stop}")
        #         start = stop + 1
        #         naac_ws[f"A{start}"].alignment = Alignment(horizontal="center", 
        #                                                    vertical="center")
        #         old_number = number
        #         # stop += 1
        #         naac_ws[f"A{start}"] = self.categories[number]
        #         width = max(width, len(self.categories[number])*1.3)
        #     stop += 1
        #     naac_ws[f"B{stop}"] = spec
        #     naac_ws[f"C{stop}"] = spec_data[spec]
        # naac_ws.column_dimensions["A"].width = width
    
        # * Entering all specification codes.
        naac_ws.append(["Classification", "Code", "Count"])
        start, word = 1, None
        for num, (classification, category) in enumerate(self.classification_list.items(), 2):
            naac_ws.append([category, classification, spec_data.get(classification, 0)])
            width = max(width, len(category)*1.2)
            if word != category:
                naac_ws.merge_cells(f"A{start}:A{num-1}")
                naac_ws[f"A{start}"].alignment = Alignment(horizontal="center", 
                                                           vertical="center")
                start, word = num, category
        else:
            naac_ws.merge_cells(f"A{start}:A{num}")
            naac_ws[f"A{start}"].alignment = Alignment(horizontal="center", 
                                                        vertical="center")
            for i in range(1, 4):
                naac_ws[f"{get_column_letter(i)}1"].alignment = Alignment(
                    horizontal="center", vertical="center")
                naac_ws[f"{get_column_letter(i)}1"].font = Font(bold=True, size=12)
            naac_ws.column_dimensions["A"].width = width

        while True:
            try:
                naac_wb.save("data/naac.xlsx")
                break
            except PermissionError:
                print("Failed to save naac.xlsx")
                ossystem("taskkill /im EXCEL.EXE naac.xlsx")
        # ossystem("start EXCEL.EXE naac.xlsx")


class DriveReader():
    """This project aims to read files in a drive and categorize them."""

    def __init__(self) -> None:
        """Initialize the class."""
        self.creds = None
        self.initialize_connection()

    def initialize_connection(self):
        """Make the initial connection with drive."""
        # The file stores user's access and refresh tokens, and is created 
        # automatically when first authorization flow is completed.
        if path.exists("token.json"):
            self.creds = Credentials.from_authorized_user_file("token.json", 
                                                               SCOPES)

        if not self.creds or not self.creds.valid:
            if self.creds and self.creds.expired and self.creds.refresh_token:
                self.creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    "credentials.json", SCOPES)
                self.creds = flow.run_local_server(port=0)
            # Save the credentials for the next run.
            with open("data/token.json", "w") as token:
                token.write(self.creds.to_json())

        # Create a connection with drive.
        self.service = build("drive", "v3", credentials=self.creds)

        if self.creds and self.creds.valid:
            return "Connection made."
        else:
            return "Connection failed."

    def search_file(self, file_name: str):
        """Search for a specific file."""
        try:
            response = self.service.files().list(
                q=f"name contains '{file_name}'"
            ).execute()
            return response.get("files", None)

        except HttpError as error:
            print(f"An error occurred: {error}")
            return None

    def search_folder(self, category_name: str):
        """Search for a specific folder."""
        try:
            response = self.service.files().list(
                q=f"name contains '{category_name}' and mimeType = \
                    'application/vnd.google-apps.folder'"
            ).execute()
            folders = response.get("files", None)
            if len(folders) > 0:
                return folders[0]
            else: return None

        except HttpError as error:
            print(f"An error occurred: {error}")
            return None

    def download_sheet(self):
        """Download the required excel sheet."""
        self.excel_sheet_id = "1b5yJfOIWCHXdr7VbFxoNLs_SI5zPR7CL0MsCI1zqWaM"
        try:
            mime_type = "application/vnd.openxmlformats-officedocument"
            mime_type += ".spreadsheetml.sheet"
            request = self.service.files().export_media(fileId=self.excel_sheet_id,
                                                        mimeType=mime_type)
            file = BytesIO()
            downloader = MediaIoBaseDownload(file, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()

            with open("data/doc_classification.xlsx", "wb") as write_file:
                write_file.write(file.getbuffer())

            print(F'Download {int(status.progress() * 100)}.')
            return True

        except HttpError as error:
            print(f"{error} has occurred.")
            return False
        
    def categorize_files(self):
        """Categorize the files in the various folders according to code."""
        try:
            with open("data/folders.json", "r") as file:
                folder_names: list[str] = load(file)
        except FileNotFoundError:
            print("Please specify the folders to search in `folders.json`.")
            return

        self.data: dict[Category, dict[Year, dict[Code, int]]] = {i: {"0": {"0": 0}} for i in self.categories}
        self.exempt: list[tuple[Name, str]] = []

        for folder_search in folder_names:
            folder = self.search_folder(folder_search)
            folder_id, folder_name = folder.get("id"), folder.get("name")
            try:
                page_token = None
                while True:
                    # Search for all files with the folder as parent.
                    response = self.service.files().list(
                        q=f"'{folder_id}' in parents and trashed = false",
                        spaces='drive',
                        fields='nextPageToken, files(name)',
                        pageToken=page_token
                    ).execute()

                    for file in response.get("files"):
                        # print(file)
                        if file.get("name") is not None:
                            if_failed = self.classify_file(file.get("name"))
                            if if_failed:
                                self.exempt.append((if_failed, folder_name))
                    page_token = response.get("nextPageToken", None)

                    if page_token is None:
                        break

            except HttpError as error:
                print(f"An error occurred: {error}")

        else:
            for category in list(self.data.keys()):
                for year, year_data in self.data[category].items():
                    self.data[category][year] = self.sort_dictionary(year_data)
                self.data[category] = self.sort_dictionary(self.data[category], True)
                if "0" in self.data[category]:
                    self.data.pop(category)

    def classify_file(self, name:str):
        """Classify the file in categories based on naming structure."""
        try:
            date, code, extra = name.split("_", 2)
            code = code.upper()
            if code not in self.code_list:
                raise KeyError
        except ValueError:
            return name
        except KeyError:
            return name
        else:
            try:
                year, month = int(date[:4]), int(date[4:6])
                # print(year, month, day)
                if month > 0 and month < 5:
                    year = f"{year-1}-{year}"
                else:
                    year = f"{year}-{year+1}"
            except ValueError:
                return name
            else:
                category = self.code_list[code][1]
                category_data = self.data.get(category, {"0": {"0": 0}})
                year_data = category_data.get(year, {"0": 0})
                code_val = year_data.get(code, 0)
                year_data.update({code: code_val + 1})
                if "0" in year_data:
                    year_data.pop("0")
                category_data.update({year: year_data})
                if "0" in category_data: 
                    category_data.pop("0")
                self.data.update({category: category_data})
                if "0" in self.data:
                    self.data.pop("0")

    def sort_dictionary(self, unsorted_dict: dict[str, ], reverse=False):
        """A util function to sort the keys of a dictionary."""
        order_list: list[str] = sorted(unsorted_dict.keys(), reverse=reverse)
        sorted_dictionary = dict({i: unsorted_dict[i] for i in order_list})
        return sorted_dictionary

    def main(self):
        """The main function of DriveReader class."""
        # self.download_sheet()
        self.excelWorker = ExcelWorker()
        self.code_list = self.excelWorker.code_list
        self.categories = self.excelWorker.classification_list.values()
        self.categorize_files()
        with open("data/data.json", "w") as file:
            data_obj = dumps(self.data, indent=4)
            file.write(data_obj)
        with open("data/exempt.json", "w") as file:
            exempt_obj = dumps(self.exempt, indent=4)
            file.write(exempt_obj)
        self.excelWorker.write_data_to_excel(self.data, self.exempt)
        self.excelWorker.write_naac_data_to_excel(self.data)


if __name__ == "__main__":
    # Driver Code
    try:
        DR = DriveReader()
        if DR.creds and DR.creds.valid:
            DR.main()
        else:
            print("Could not run the program due to invalid credentials.")
            print("Fix credentials and try again.")
            sysexit()
    except KeyboardInterrupt:
        print("\n\nExiting program by interrupt.")